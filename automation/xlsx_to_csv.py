#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# 

import sys
import openpyxl
import csv
import chardet



def safeint(s):
    if not s:
        return s
    f = float(s)
    r = round(f)
    if f == r:
        return int(f)
    else:
        raise ValueError(f"Can't parse {s} as int without losing precision")


def represents_int(s):
    try:
        safeint(s)
        return True
    except (ValueError, TypeError):
        return False


def represents_float(s):
    try:
        float(s)
        return True
    except (ValueError, TypeError):
        return False


def rows_from_xlsx(filename, header_row=0, sheet_index=0, sheet_name=None, skip_rows=1):
    # Öffne die Excel-Datei mit openpyxl
    workbook = openpyxl.load_workbook(filename, data_only=True)

    # Wähle das gewünschte Arbeitsblatt aus
    if sheet_name:
        sheet = workbook[sheet_name]
    else:
        sheet = workbook.worksheets[sheet_index]

    rows = []

    # Überspringe die ersten Zeilen (header_row + skip_rows)
    for _ in range(header_row + skip_rows):
        next(sheet.iter_rows())

    # Lese die Spaltenüberschriften (Header)
    headers = [cell.value for cell in next(sheet.iter_rows())]

    # Iteriere über die Datenzeilen und erstelle ein Wörterbuch für jede Zeile
    for row in sheet.iter_rows(values_only=True):
        # Verwende die zip-Funktion, um Header und Werte zu verknüpfen und ein Wörterbuch zu erstellen
        row_data = dict(zip(headers, row))
        rows.append(row_data)

    return rows


def parse_xlsx(book, header_row=0, sheet_index=0, sheet_name=None, skip_rows=1):
    rows = []
    
    # Wähle das gewünschte Arbeitsblatt aus
    if sheet_name:
        sheet = book[sheet_name]
    else:
        sheet = book.worksheets[sheet_index]
    
    # Bestimme die Anzahl der Spalten (ncols)
    ncols = sheet.max_column

    # Erzeuge die Spaltenüberschriften
    headers = {c: sheet.cell(row=header_row + 1, column=c + 1).value or f"{openpyxl.utils.get_column_letter(c + 1)}" for c in range(ncols)}

    for row_num in range(header_row + skip_rows + 1, sheet.max_row + 1):
        entry = {}
        for col_num, header in headers.items():
            cell = sheet.cell(row=row_num, column=col_num + 1)
            cell_type = cell.data_type
            value = cell.value
            
            if cell_type == "n" and cell.is_date:
                entry[header] = cell.value
            elif cell_type == "n":
                if int(value) == value:
                    entry[header] = int(value)
                else:
                    entry[header] = float(value)
            elif cell_type == "s":
                entry[header] = value
            elif cell_type == "b":
                entry[header] = bool(value)
            else:
                entry[header] = None

        rows.append(entry)

    return rows


def rows_from_csv(input_file, delimiter=';', input_encoding='detect'):
    if input_encoding == 'detect':
        input_encoding = _detect_file_encoding(input_file)

    with open(input_file, 'r', encoding=input_encoding) as f:
        # if a list of delimiters is given, try to sniff the correct one
        if isinstance(delimiter, list):
            dialect = csv.Sniffer().sniff(f.read(), delimiter)
            f.seek(0)
            reader = csv.DictReader(f, dialect=dialect)
        else:
            reader = csv.DictReader(f, delimiter=delimiter)
        rows = [r for r in reader]
    return rows


def _detect_file_encoding(input_file, line_count=20):
    with open(input_file, 'rb') as f:
        # read some lines
         rawdata = b''.join([f.readline() for _ in range(line_count)])
    return chardet.detect(rawdata)['encoding']


def write_csv(rows, out=None, skip_header=False):
    header = rows[0].keys()
    outfile = open(out, 'w', newline='') if out else sys.stdout

    writer = csv.DictWriter(
        outfile,
        header,
        delimiter=',',
        quotechar='"',
        lineterminator='\n',
        quoting=csv.QUOTE_MINIMAL
    )
    if not skip_header:
        writer.writeheader()
    writer.writerows(rows)

    if outfile is not sys.stdout:
       outfile.close()