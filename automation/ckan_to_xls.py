# -*- coding: utf-8 -*-
"""Extract CKAN metadaten to a XLS file

Usage:
  ckan_to_xls.py --dataset <dataset-name> --file <path-to-file> [--no-verify]
  ckan_to_xls.py (-h | --help)
  ckan_to_xls.py --version

Options:
  -h, --help                   Show this screen.
  --version                    Show version.
  -d, --dataset <dataset-name> Name of the dataset to extract the metadata.
  -f, --file <path-to-file>    Path to XLS file
  --no-verify                  Option to disable SSL verification for requests.

"""

import os
import sys
import traceback
import re
import json
import collections
from copy import copy
import openpyxl
from docopt import docopt
from ckanapi import RemoteCKAN, NotFound
from dotenv import load_dotenv, find_dotenv

__location__ = os.path.realpath(
    os.path.join(
        os.getcwd(),
        os.path.dirname(__file__)
    )
)

load_dotenv(find_dotenv())
arguments = docopt(__doc__, version='Extract metadata to XLS 1.0')


def map_metadata_to_xls(metadata):
    return {
        'D2': metadata['title'],
	'D3': metadata['notes'],
        'D4': ", ".join([g['title'] for g in metadata['groups']]),
        'D5': metadata['legalInformation'],
        'D6': metadata['spatialRelationship'],
        'D7': metadata['url'],
        'D8': metadata['author'],
        'D9': metadata['timeRange'],
        'D10': metadata['dataQuality'],
        'D11': metadata['dateFirstPublished'],
        'D12': metadata['dateLastUpdated'],
        'D13': metadata['dataType'][0],
        'D14': metadata['updateInterval'][0],
        'D15': ", ".join([t['name'] for t in metadata['tags']]),
        'D16': metadata['version'],
        'D17': metadata['license_id'],
    }

def convert_comments(comments):
    if not comments:
       return {}

    result = collections.defaultdict(str)

    lines = comments.split('\n')
    key = None
    for line in lines:
        line = line.strip()
        m = re.match(r'^\*\*(.*?)\*\*$', line)
        if m:
            key = m[1]
            continue
        if key:
            if line or result[key]:
                result[key] += f"{line}\n"
	
    return result

def copy_style(source, target):
    if source.has_style:
        target.font = copy(source.font)
        target.border = copy(source.border)
        target.fill = copy(source.fill)
        target.number_format = copy(source.number_format)
        target.protection = copy(source.protection)
        target.alignment = copy(source.alignment)

try:
    BASE_URL = os.getenv('CKAN_BASE_URL')
    API_KEY = os.getenv('CKAN_API_KEY')
    ckan = RemoteCKAN(BASE_URL, apikey=API_KEY)


    # CKAN metadata
    dataset = arguments['--dataset']
    data = {
        'id': dataset,
    }
    print(f"Extracting metadata from dataset {dataset}.")
    try:
        if arguments['--no-verify']:
            ckan_metadata = ckan.call_action('package_show', data, requests_kwargs={'verify': False})
        else:
            ckan_metadata = ckan.call_action('package_show', data)
    except NotFound:
         print('Dataset %s not found!' % dataset, file=sys.stderr)
         raise
    
    # XLS file
    template_path = os.path.join(__location__, 'OGD-Metadaten_Template.xlsx')
    wb = openpyxl.load_workbook(template_path)
    ws = wb['metadata']
    # set all values
    xls_metadata = map_metadata_to_xls(ckan_metadata)
    for cell, value in xls_metadata.items():
        ws[cell] = value
    

    # insert bemerkungen
    comments = convert_comments(ckan_metadata['sszBemerkungen'])
    comment_row = 21 
    label_cell = ws.cell(column=1, row=comment_row)
    content_cell = ws.cell(column=2, row=comment_row)
    for title, comment in comments.items():
        ws.insert_rows(comment_row)
        label = ws.cell(column=1, row=comment_row, value='bemerkung') 
        title = ws.cell(column=2, row=comment_row, value=title) 
        content = ws.cell(column=3, row=comment_row, value=comment) 
        
        copy_style(label_cell, label)
        copy_style(content_cell, title)
        copy_style(content_cell, content)
        
        comment_row += 1

    # find row of attributes
    attributes_row = None
    for row in ws.iter_rows(min_row=comment_row, max_row=ws.max_row, min_col=1, max_col=1):
        if row[0].value == 'attributelement':
            attributes_row = row[0].row
            break
    # insert attributes
    attributes = json.loads(ckan_metadata['sszFields'])
    for attribute in attributes:
        m = re.match(r'^(?P<name>.+?)(\s*\(technisch: (?P<tech_name>.+?)\))?$', attribute[0])
        assert m, f"Could not match name and tech_name from {attribute[0]}"
        desc = attribute[1]
        ws.insert_rows(attributes_row)
        label = ws.cell(column=1, row=attributes_row, value='attributelement') 
        tech_name = ws.cell(column=2, row=attributes_row, value=m['tech_name']) 
        name = ws.cell(column=3, row=attributes_row, value=m['name']) 
        description = ws.cell(column=4, row=attributes_row, value=desc) 
        
        copy_style(label_cell, label)
        copy_style(content_cell, tech_name)
        copy_style(content_cell, name)
        copy_style(content_cell, description)
        
        attributes_row += 1

    # save the file
    xls_path = arguments['--file']
    wb.save(xls_path)

    print(f"Saved metadata to {xls_path}.")
    
except Exception as e:
    print("Error: %s" % e, file=sys.stderr)
    print(traceback.format_exc(), file=sys.stderr)
    sys.exit(1)
